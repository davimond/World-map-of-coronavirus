###
### 从国内网站爬取新冠病毒最新数据，用颜色标注在世界地图中。
### 世界地图具有地图热点功能
###


import time
CheckPoint = time.time()
print("加载第三方package... ",end = '')

from pyecharts import options as opts
from pyecharts.charts import Map
from pyecharts.globals import ThemeType
from pyecharts.globals import JsCode
import pprint
import requests          #reauests 用于爬取的数据,获取网页数据
import io
import sys
import json
import openpyxl

def print_elaspe_time(lasttime):
    print(str(round(time.time()-lasttime,4)) + '秒')

print_elaspe_time(CheckPoint)
CheckPoint = time.time()

sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')
heads = {}
###模仿浏览器访问网站行为，避免被误判为爬虫。真的这么简单就骗过去了吗?
heads['User-Agent'] = 'Mozilla/5.0 ' \
                          '(Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 ' \
                          '(KHTML, like Gecko) Version/5.1 Safari/534.50'

#❖❖❖ Step 1:  从网页获取中国、各国新冠病毒确诊累计总数。
#---------------------------------------------------------------------
### 不同的网站取数据方法有差异，按网站单独处理。
web_site = "丁香园网站"
web_site = "百度网站"

print("获取" + web_site + '数据... ',end = '')
if web_site == "丁香园网站":
    #################################从丁香园网站爬数据############################################
    # 爬取的网页URL地址
    url = "https://ncov.dxy.cn/ncovh5/view/pneumonia?from=groupmessage&isappinstalled=0"
    response = requests.get(url)
    # response.encoding = 'utf-8'
    response.encoding = None
    #### 网页内容保存在result字符串中
    result = response.text
    ### 网页内容写入dingxiang.txt 文件（临时分析用）
    with open("dingxiang.txt",'w') as f:
        f.write(result)
    ### 设置两个关键字，并从中匹配出全世界各国的新冠病毒数据
    match_string1 = "window.getListByCountryTypeService2true = "
    match_string2 = "catch(e){}"
    pos1 = result.index(match_string1) + 42
    ## match_string2 关键字在网页中出现3处，指定搜索的起始位置，
    ## 以避免前面出现的相同关键字造成搜索干扰。
    pos2 = result.index(match_string2,pos1) - 1
    #
    #截取出各国的字符串数据放入world_string，外层列表，其中元素为字典 [{},{}...{}]
    world_string = result[pos1:pos2]
    ### 将截取的字符串写入本地文件（临时分析使用）
    with open("cutstring.txt","w") as f:
        f.write(world_string)
    #
    ### 将字符串数据转为列表数据，保存在world_data中。
    world_data = json.loads(world_string)
    #
    ####获取全国新冠肺炎累计确认数据
    ### 从字符串countRemark开始查找数据
    match_string = 'countRemark'
    pos = result.index(match_string)
    pos1 = pos + 15
    pos2 = pos1 + 100
    #
    ### 从countRemark关键字后的15位开始，截取100个字符。缩小搜索范围。
    ### 其中"confirmedCount"即为新冠肺炎累计确认数据
    cut_string = result[pos1:pos2]
    #
    pos1 = cut_string.index('confirmedCount') + 16
    pos2 = cut_string.index('suspectedCount') - 2
    ### 中国冠状病毒确诊累计总数
    ChinaCounts = int(float(cut_string[pos1:pos2]))
    #### 从网页获取数据完毕，且将数据转换到列表中
    ############################丁香园网站获取数据完毕#######################
    #####################################################################
elif web_site == "百度网站":
    ##########################从百度网页取数据######################
    # 爬取的网页URL地址
    url = "https://voice.baidu.com/act/newpneumonia/newpneumonia?from=groupmessage&isappinstalled=0"
    response = requests.get(url)
    response.encoding = None
    response.encoding = 'utf-8'
    #### 网页内容保存在result字符串中
    result = response.text
    with open("baidu.txt",'w') as f:
        f.write(result)             #写入文件，供临时分析使用
    #
    ### 获取世界各国冠状病毒数据。在web页面中国和外国数据放在不同的数据结构中，需要单独解析数据。
    ### 设置两个搜索关键字:caseOutsideList，dataSource.并从中匹配出全世界各国的新冠病毒数据
    ### 两个关键字之间的内容：
    ### "caseOutsideList":[
    ### {"confirmed":"1","died":"","crued":"","icuDisable":"1","area":"\u4f0a\u62c9\u514b","subList":[]},
    ### {"confirmed":"1","died":"","crued":"","icuDisable":"1","area":"\u9ece\u5df4\u5ae9","subList":[]},
    ###{"confirmed":"1","died":"","crued":"","icuDisable":"1","area":"\u4ee5\u8272\u5217","subList":[]},
    ###  .............
    ### "dataSource":"\u6570\u636e"
    #
    match_string1 = "caseOutsideList"
    match_string2 = 'dataSource'
    pos1 = result.index(match_string1)+17
    pos2 = result.index(match_string2) - 2
    #
    #截取出各国的字符串数据放入world_string，外层列表，其中元素为字典 [{},{}...{}]
    world_string = result[pos1:pos2]
    ### 将网页数据写入本地文件
    with open("cutstring.txt","w") as f:
        f.write(world_string)           #写入文件，供临时分析使用
    #
    ### 将字符串数据转为列表数据，保存在world_data中。
    world_data = json.loads(world_string)
    #
    ### 解析网页内容，单独获取中国的冠状病毒确诊累计总数
    ### 百度网数据的关键字，从summaryDataIn开始查找数据。同样先截取60个字符后，缩小匹配字符串搜索的范围。
    ### 截取60个字符，内容如下：
    ### "summaryDataIn":{"confirmed":"77048","died":"2445","cured":"23171","unconfirmed":"4148","relativeTime"
    ### 两个关键字：summaryDataIN、died
    match_string = 'summaryDataIn'
    pos = result.index(match_string)
    pos1 = pos + 28
    pos2 = pos + 60
    cut_string = result[pos1:pos2]
    #
    post2 = cut_string.index("died")
    ### 中国冠状病毒确诊累计总数
    ChinaCounts = int(float(cut_string[1:post2 - 3]))
    ############### 百度网站获取数据完毕 ##################
    #### 从网页获取数据完毕，且将数据转换保存到列表world_data，和变量ChinaCounts中
else:
    print("未指定网站，或不支持该网站数据的解析")
    print("网站名称： ",web_site)
    sys.exit(0)

print_elaspe_time(CheckPoint)
CheckPoint = time.time()
print('转换Excel国家代码... ',end = '')


#❖❖❖ Step 2:  读取Excel文件，获取4列数据：2位国家代码、3位国家代码，中英文国家名称。
#----------------------------------------------------------------------

excel = openpyxl.load_workbook('国家代码23.xlsx')
#获取第一个表单，序号0
mysheet = excel.worksheets[0]
#### 最大行数：mysheet.max_row , 最大列数：mysheet.max_column

#列表用于存储2位国家名称代码，国名，ConvertTable转换表供JavaScript使用
myCountries =[]
ConvertTable_Country ='    var names={'
ConvertTable_2code   ='    var flags={'
First = True

### 列表myCountries,元素为字典，格式：
### myCountries[ {'CountryName': '安道尔', 'Ename':'Andorra','Code2': 'AD', 'Code3': 'AND'},.......]
#取表格A2:C248区域，列字段分别为2字母国家代码，3字母国家代码，国家名称
for rowObject in mysheet['A2':'F'+str(mysheet.max_row)]:
    ## print(rowObject[0].value,rowObject[1].value,rowObject[2].value)
    myCountries.append({"CountryName":rowObject[2].value,
                        "Ename":rowObject[5].value,
                        "Code2":rowObject[0].value.lower(),"Code3":rowObject[1].value})
    #--生成英文-中文国家名称代码表；英文国名-2code转换表，供javascript 使用------
    if First:
        comma =''
        First = False
    else:
        comma =','
    ConvertTable_2code += comma + "'"+rowObject[5].value+"':'" + rowObject[0].value.lower() + ".png'"
    ConvertTable_Country += comma + "'"+rowObject[5].value+"':'" + rowObject[2].value + "'"

ConvertTable_2code += '};\n'
ConvertTable_Country += '};\n'
with open('国家代码23.txt','w') as f:  # 保存国家名代码对照表，供其他项目使用
    f.write('MyContries = '+pprint.pformat(myCountries,width=300))

print_elaspe_time(CheckPoint)
CheckPoint = time.time()
print('生成各国对应数据... ',end = '')

#❖❖❖ Step 3: 生成2位国家代码的病毒数据
#---------------------------------------------------------------------
#### 根据Step1生成的各国病毒数据的列表：world_data、国家代码的数据字典：myCountries,
#### 生成2位国家代码、病例确诊数对照数据all_coronavirus[ ['jp',500],['us':33].....]

all_coronavirus = []

### 从百度和丁香园生成的world_data的数据结构有差异，而且数据字典的Key值不同需分别处理。
if web_site == "百度网站":
    ### 以myCountries{} 为外循环
    for data1 in myCountries:
        ## 列表myCountries,元素为字典，格式：
        ### myCountries = [{'CountryName': '安道尔', 'Ename':'Andorra','Code2': 'AD', 'Code3': 'AND'},.......]
        ## world_data 格式为列表嵌套字典 [{"area":"意大利",.....,"confirmed":73,}.....]
        ## world_data 与 myCountries 用国家名字（汉字）匹配，获取英文国家名称Ename
        if data1["CountryName"] == '中国':
            all_coronavirus.append(['China', ChinaCounts])
        else:
            findit = False
            for data2 in world_data:
                if data1["CountryName"] == data2["area"]:
                    all_coronavirus.append([data1["Ename"], int(data2["confirmed"])])
                    findit = True
                    break
            if not findit:
                ## 未能匹配国家名称，国名名称不在网站数据列表中....
                all_coronavirus.append([data1["Ename"],0])

if web_site == "丁香园网站":
    ### 以myCountries{} 为外循环
    for data1 in myCountries:
        ## 列表myCountries,元素为字典，格式：
        ### myCountries = [{'CountryName': '安道尔', 'Ename':'Andorra','Code2': 'AD', 'Code3': 'AND'},.......]
        ## world_data 格式为列表嵌套字典[{"provinceName":"意大利",.....,"confirmedCount":73,}....]
        ## world_data 与 myCountries 匹配国家名字（汉字），获取2字母国家代码
        if data1['CountryName'] == '中国':
            all_coronavirus.append([data1['Ename'],ChinaCounts])
        else:
            findit = False
            for data2 in world_data:
                if data1['CountryName'] == data2['provinceName']:
                    all_coronavirus.append([data1['Ename'],int(data2['confirmedCount'])])
                    findit = True
                    break
            if not findit:
                ## 未能匹配国家名称，国名名称不在网站数据列表中....
                all_coronavirus.append([data1['Ename'],0])


print_elaspe_time(CheckPoint)
CheckPoint = time.time()
print('绘制病毒世界地图... ',end = '')

js_code1 =  """
function (params) {
    var result = '';
    var Ename = params.name; 
    var virus = params.value;
"""

js_code2 =  """
    var flag = flags[Ename];
    var Country = names[Ename];
    result += '<img src="./CountryFlag/'+ flag +'" height=25 /> ' + Ename  + '<br/>';
    result += Country + ' ✦ 确诊数: ' + virus + '<img src="./CountryFlag/coronavirus.gif" height=20 /> ';
    console.log(params);
    return result
    }
"""
js_code = js_code1 + ConvertTable_Country + ConvertTable_2code + js_code2

#❖❖❖ Step 4:  通过pyecharts Map的视觉映射功能对数据分组，以不同的颜色显示
#---------------------------------------------------------------------

curr_time = time.strftime("%H:%M")   ##24小时格式
curr_date = time.strftime("%m-%d")

virusmap = Map(opts.InitOpts(width='1500px',height='1200px',
    page_title="新冠病毒世界地图",theme=ThemeType.ROMANTIC))

virusmap.add("累计确诊病毒人数",all_coronavirus,"world",is_map_symbol_show=False)
    ###### 设置高亮地图区域颜色
    #emphasis_itemstyle_opts = opts.ItemStyleOpts(opacity=0.5, area_color='brown’),
    ###### 设置高亮地图区域Label（地名）的颜色
    #emphasis_label_opts = opts.ItemStyleOpts(color='yellow'),

virusmap.set_global_opts(
    ## 对整个图形的抬头标题进行设置
    title_opts=opts.TitleOpts( title='全球新冠病毒地图',
        title_textstyle_opts=opts.TextStyleOpts(font_size=35,color='brown'),
        subtitle='(数据来源:'  + web_site + ' '  + curr_time  + "@" + curr_date  + ")",
        subtitle_textstyle_opts=opts.TextStyleOpts(color='#666666'),
        pos_left='center',pos_top='50px'),
    ## 对图例进行配置，取消顶部的图例显示
    legend_opts=opts.LegendOpts(is_show=False),
    ## 配置浮动提示窗口参数，背景色、字体
    tooltip_opts=opts.TooltipOpts(  #####background_color='#83c1f5',
        textstyle_opts=opts.TextStyleOpts(font_size=20),formatter=JsCode(js_code)),
    ## 配置视觉映射参数：分段的设置、颜色、label等等
    visualmap_opts=opts.VisualMapOpts(pieces=[
        {"min":80000, 'color':'#8b0b0b'},
        {"min":40000, "max":79999,'color':'#d34833'},
        {"min":20000, "max":39999,'color':'#d5694e'}, #d5694e
        {"min":10000, "max":19999},
        {"min":5000,  "max":9999},
        {"min":2000,  "max":4999},
        {"min":1000,  "max":1999},
        {"min":500,   "max":999},
        {"min":200,   "max":499},
        {"min":100,   "max":199},
        {"min":50,    "max":99},
        {"min":1,     "max":49},
        {"value":0,'color':'#dcdcdc'}],
        pos_bottom='35%',pos_left='130px',
        is_piecewise=True,
        textstyle_opts=opts.TextStyleOpts(color='#888888'))
)

### is_show=False,不显示Series，即在地图中不显示国名，鼠标覆盖才显示
virusmap.set_series_opts(label_opts=opts.LabelOpts(is_show=False))
virusmap.render("NewVirusMap.html")


print_elaspe_time(CheckPoint)
#print("\n未感染病毒的国家数数量：",len(group0))
#print("感染病毒的国家数数量：",len(group1) + len(group2) + len(group3) + len(group4) + len(group5) -1)
print("=== 程序运行结束 ===")



